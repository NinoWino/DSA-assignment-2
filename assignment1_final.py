import re
import pickle
import hashlib
from openpyxl import Workbook, load_workbook
from colorama import Fore, Style, init
import logging
from datetime import datetime, timedelta, timezone
import heapq
import itertools
import json
from collections import Counter
import random
from models import Student, StudentRequest, RequestQueue, StudentBST
from graphviz import Digraph
import os
from dotenv import load_dotenv
import requests
import time
import pandas as pd
import matplotlib
matplotlib.use("Agg")  # headless-safe
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from datetime import datetime

load_dotenv()

init(autoreset=True)

logging.basicConfig(
    filename='student_system.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

STORAGE_FILE = "student_data.pkl"
student_tree = StudentBST()

REQUEST_FILE = "requests_data.json"

def save_requests():
    with open(REQUEST_FILE, "w") as f:
        f.write(request_queue.to_json())
    logging.info("Request queue saved.")

def load_requests():
    global request_queue
    if os.path.exists(REQUEST_FILE):
        with open(REQUEST_FILE) as f:
            request_queue = RequestQueue.from_json(f.read())
        logging.info("Request queue loaded.")


def save_data():
    with open(STORAGE_FILE, 'wb') as f:
        pickle.dump(student_tree, f)
    logging.info("Student data saved to persistent storage.")

def load_data():
    global student_tree
    if not os.path.exists(STORAGE_FILE):
        return

    with open(STORAGE_FILE, 'rb') as f:
        data = pickle.load(f)

    # if it’s still the old dict format, re-build the BST
    if isinstance(data, dict):
        old_dict = data
        bst = StudentBST()
        for student in old_dict.values():
            bst.insert(student)
        student_tree = bst
        logging.info("Migrated old dict data into StudentBST.")
    # if it’s already a BST, just assign it
    elif isinstance(data, StudentBST):
        student_tree = data
        logging.info("Loaded StudentBST from storage.")
    else:
        raise RuntimeError(f"Unexpected data type in {STORAGE_FILE}: {type(data)}")

def valid_email_format(email):
    return re.match(r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$", email)

def valid_course_code(course):
    return re.match(r"^[A-Z]{2,4}\d{3}$", course)

def display_all_students():
    any_printed = False
    for student in student_tree.in_order_traversal():
        student.display_details()
        any_printed = True
    if not any_printed:
        print("No students registered.")

def add_student():
    try:
        student_id = int(input("What is the student ID: "))
        if student_tree.search(student_id) is not None:
            print("Student ID already registered.")
            return

        name = input("What is the student name: ").strip()
        email = input("What is the student email: ").strip()
        if not valid_email_format(email):
            print("Invalid email format.")
            return

        course_list = input("Enter student course codes (e.g., CS123 IT345 ITX678): ").split()
        course_list = [course.strip().upper() for course in course_list]
        for course in course_list:
            if not valid_course_code(course):
                print(f"Invalid course code format: {course}")
                return

        year_of_study = int(input("What year of study is the student in (1-3): "))
        if year_of_study not in [1, 2, 3]:
            print("Enter a valid year of study (1, 2, or 3).")
            return

        full_time_status = input("Is the student full-time? (YES/NO): ").strip().upper()
        if full_time_status == "YES":
            full_time = True
        elif full_time_status == "NO":
            full_time = False
        else:
            print("Please enter YES or NO.")
            return

        new_student = Student(
            name, student_id, email, course_list, year_of_study, full_time
        )
        student_tree.insert(new_student)
        logging.info(f"Student {student_id} - {name} added.")
        save_data()
        print("Student successfully added.")

    except Exception as e:
        print("Error: Invalid format")
        print(str(e))

def enroll_course():
    """
    Prompt for a student ID and a course code, validate both,
    then enroll the student in the course (if not already enrolled).
    """
    try:
        stud_id = int(input("Enter student ID: ").strip())
    except ValueError:
        print("Invalid input. Student ID must be a number.")
        return

    student = student_tree.search(stud_id)
    if not student:
        print("Student ID not found.")
        return

    course = input("Enter course code to enroll: ").strip().upper()
    if not valid_course_code(course):
        print("Invalid course code format. Format must be like CS123.")
        return

    if course in student.course_list:
        print(f"Student {stud_id} is already enrolled in {course}.")
        return

    student.add_course(course)
    save_data()
    logging.info(f"Student {stud_id} enrolled in course {course}.")
    print(f"Course {course} successfully added to student {stud_id}.")

def remove_student_course():
    """
    Prompt for a student ID and a course code, validate both,
    then remove the course from the student's list (if present).
    """
    try:
        stud_id = int(input("Enter student ID: ").strip())
    except ValueError:
        print("Invalid input. Student ID must be a number.")
        return

    student = student_tree.search(stud_id)
    if not student:
        print("Student ID not found.")
        return

    course = input("Enter course code to remove: ").strip().upper()
    if not valid_course_code(course):
        print("Invalid course code format. Format must be like CS123.")
        return

    if course not in student.course_list:
        print(f"Course {course} not found for student {stud_id}.")
        return

    student.remove_course(course)
    save_data()
    logging.info(f"Course {course} removed from student {stud_id}.")
    print(f"Course {course} successfully removed from student {stud_id}.")

def bubble_sort_year_of_study():
    sorted_list = list(student_tree.values())
    for i in range(len(sorted_list)):
        for j in range(0, len(sorted_list) - i - 1):
            if sorted_list[j].year_of_study > sorted_list[j + 1].year_of_study:
                sorted_list[j], sorted_list[j + 1] = sorted_list[j + 1], sorted_list[j]
    print("Students sorted by year of study (ascending):")
    for s in sorted_list:
        s.display_details()

def selection_sort_num_reg_course():
    sorted_list = list(student_tree.values())
    n = len(sorted_list)
    for i in range(n):
        max_index = i
        for j in range(i + 1, n):
            if len(sorted_list[j].course_list) > len(sorted_list[max_index].course_list):
                max_index = j
        sorted_list[i], sorted_list[max_index] = sorted_list[max_index], sorted_list[i]
    print("Students sorted by number of registered courses (descending):")
    for s in sorted_list:
        s.display_details()

def search_student():
    key = input("Enter student ID or Name to search: ").strip()
    logging.info(f"Search performed for student: {key}")
    found = False

    try:
        key_id = int(key)
        student = student_tree.search(key_id)
        if student:
            student.display_details()
            return
    except ValueError:
        pass

    for student in student_tree.in_order_traversal():
        if student.name.lower() == key.lower():
            student.display_details()
            found = True

    if not found:
        print("Student not found.")

def export_to_excel():
    if not student_tree:
        print("No students to export.")
        return

    filename = input("Enter a name for the Excel file (without extension): ").strip()
    if not filename.lower().endswith(".xlsx"):
        filename += ".xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    ws.append(["Student ID", "Name", "Email", "Courses", "Year of Study", "Full-time"])

    for student in student_tree.values():
        ws.append([
            student.student_id,
            student.name,
            student.email,
            ', '.join(student.course_list),
            student.year_of_study,
            "Yes" if student.is_full_time else "No"
        ])

    try:
        wb.save(filename)
        print(f"Student data successfully exported to '{filename}'")
    except Exception as e:
        print("Failed to save Excel file:", e)

def import_from_excel():
    """
    Prompt for an Excel file, read each row as a Student,
    and insert or update into our BST (student_tree).
    """
    filename = input("Enter the Excel filename to import (e.g., student_data.xlsx): ").strip()
    try:
        wb = load_workbook(filename)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            student_id    = int(row[0])
            name          = row[1]
            email         = row[2]
            course_list   = [c.strip().upper() for c in row[3].split(',')] if row[3] else []
            year_of_study = int(row[4])
            is_full_time  = row[5].strip().lower() == "yes"

            # Build a Student object
            new_student = Student(
                name, student_id, email,
                course_list, year_of_study, is_full_time
            )

            existing = student_tree.search(student_id)
            if existing:
                # update existing student
                existing.name          = name
                existing.email         = email
                existing.course_list   = course_list
                existing.year_of_study = year_of_study
                existing.is_full_time  = is_full_time
                logging.info(f"Updated student {student_id} from Excel import.")
            else:
                # insert a new student node
                student_tree.insert(new_student)
                logging.info(f"Imported new student {student_id} from Excel.")

        save_data()
        print(f"Student data successfully imported from '{filename}'.")
    except FileNotFoundError:
        print(f"File '{filename}' not found.")
    except Exception as e:
        print("Failed to import Excel file:", e)


def quick_sort_students(students):
    if len(students) <= 1:
        return students

    pivot = students[len(students) // 2]
    left, middle, right = [], [], []

    for s in students:
        # compare (year, name) tuples
        key = (s.year_of_study, s.name.lower())
        pivot_key = (pivot.year_of_study, pivot.name.lower())

        if key < pivot_key:
            left.append(s)
        elif key > pivot_key:
            right.append(s)
        else:
            middle.append(s)

    return quick_sort_students(left) + middle + quick_sort_students(right)


def quick_sort_year_name():
    students = list(student_tree.values())
    if not students:
        print("No students to sort.")
        return

    sorted_list = quick_sort_students(students)

    headers = ["ID", "Name", "Email", "Courses", "Year", "Full-time"]
    rows = [
        [
            str(s.student_id),
            s.name,
            s.email,
            ", ".join(s.course_list),
            str(s.year_of_study),
            "Yes" if s.is_full_time else "No"
        ]
        for s in sorted_list
    ]

    # compute column widths
    col_widths = [
        max(len(headers[i]), *(len(row[i]) for row in rows))
        for i in range(len(headers))
    ]

    # colourized header
    header_parts = [
        f"{Fore.CYAN}{Style.BRIGHT}{headers[i].ljust(col_widths[i])}{Style.RESET_ALL}"
        for i in range(len(headers))
    ]
    print(" | ".join(header_parts))
    print("-+-".join("-" * col_widths[i] for i in range(len(headers))))

    # colour rows by Year of Study: 1→red, 2→yellow, 3+→green
    for row in rows:
        yr = int(row[4])
        color = Fore.RED if yr == 1 else (Fore.YELLOW if yr == 2 else Fore.GREEN)
        line = " | ".join(row[i].ljust(col_widths[i]) for i in range(len(headers)))
        print(f"{color}{line}{Style.RESET_ALL}")


def merge_sort_students(students):
    """Recursively merge-sort by (num_courses, student_id)."""
    if len(students) <= 1:
        return students

    mid = len(students) // 2
    left = merge_sort_students(students[:mid])
    right = merge_sort_students(students[mid:])

    # merge step
    merged = []
    i = j = 0
    while i < len(left) and j < len(right):
        left_key  = (len(left[i].course_list), left[i].student_id)
        right_key = (len(right[j].course_list), right[j].student_id)

        if left_key <= right_key:
            merged.append(left[i])
            i += 1
        else:
            merged.append(right[j])
            j += 1

    # append any leftovers
    merged.extend(left[i:])
    merged.extend(right[j:])
    return merged


def merge_sort_by_courses_and_id():
    students = list(student_tree.values())
    if not students:
        print("No students in the system.")
        return

    try:
        year = int(input("Enter Year of Study to display: ").strip())
    except ValueError:
        print("Invalid year. Please enter an integer.")
        return

    filtered = [s for s in students if s.year_of_study == year]
    if not filtered:
        print(f"No students found for Year {year}.")
        return

    sorted_list = merge_sort_students(filtered)

    headers = ["ID", "Name", "Email", "Courses", "Year", "Full-time"]
    rows = [
        [
            str(s.student_id),
            s.name,
            s.email,
            ", ".join(s.course_list),
            str(s.year_of_study),
            "Yes" if s.is_full_time else "No"
        ]
        for s in sorted_list
    ]

    # compute column widths
    col_widths = [
        max(len(headers[i]), *(len(row[i]) for row in rows))
        for i in range(len(headers))
    ]

    # colourized header
    header_parts = [
        f"{Fore.CYAN}{Style.BRIGHT}{headers[i].ljust(col_widths[i])}{Style.RESET_ALL}"
        for i in range(len(headers))
    ]
    print(" | ".join(header_parts))
    print("-+-".join("-" * col_widths[i] for i in range(len(headers))))

    # (all rows are same year, but we'll still colour by that year)
    for row in rows:
        yr = int(row[4])
        color = Fore.RED if yr == 1 else (Fore.YELLOW if yr == 2 else Fore.GREEN)
        line = " | ".join(row[i].ljust(col_widths[i]) for i in range(len(headers)))
        print(f"{color}{line}{Style.RESET_ALL}")

# --- Example menu integration ---
request_queue = RequestQueue()
undo_stack = []
redo_stack = []

def undo_action():
    if not undo_stack:
        print("Nothing to undo.")
        return

    action, req = undo_stack.pop()
    if action == "enqueue":
        # undo enqueue => remove that specific request
        request_queue.remove_request(req.request_id)
        save_requests()  # ✅ persist change
        print(f"Undid enqueue of {req}")
        redo_stack.append(("enqueue", req))

    elif action == "dequeue":
        # undo dequeue => re-enqueue the same request
        request_queue.enqueue(req)
        save_requests()  # ✅ persist change
        print(f"Undid dequeue (re-enqueued) {req}")
        redo_stack.append(("dequeue", req))


def redo_action():
    if not redo_stack:
        print("Nothing to redo.")
        return

    action, req = redo_stack.pop()
    if action == "enqueue":
        # redo enqueue => put it back
        request_queue.enqueue(req)
        save_requests()  # ✅ persist change
        print(f"Redid enqueue of {req}")
        undo_stack.append(("enqueue", req))

    elif action == "dequeue":
        # redo dequeue => remove next (should be that req)
        request_queue.remove_request(req.request_id)
        save_requests()  # ✅ persist change
        print(f"Redid dequeue (removed) {req}")
        undo_stack.append(("dequeue", req))


def view_requests_menu():
    pending = request_queue.list_all()
    if not pending:
        print("No pending requests.")
    else:
        display_requests_table(pending)


def student_exists(sid):
    """
    Sequentially search through student_registration_list
    to see if a student with ID==sid is registered.
    """
    for key in student_tree.keys():
        if key == sid:
            return True
    return False

def view_queue_stats_menu():
    """
    3c. Show total, filter by type or priority, or give full breakdown.
    """
    total = request_queue.size()
    print(f"Total requests in queue: {total}\n")

    print("1. Filter by Request Type")
    print("2. Filter by Priority Level")
    print("3. Summary of all Request Types and Priority Levels")
    choice = input("Enter choice: ").strip()

    all_req = request_queue.list_all()
    if choice == '1':
        rt = input("Enter Request Type to filter: ").strip().lower()
        count = sum(1 for r in all_req if r.request_type.lower() == rt)
        print(f"{count} request(s) with type '{rt}'.")
    elif choice == '2':
        lvl_input = input("Enter Priority Level to filter: ").strip()
        try:
            lvl = int(lvl_input)
            count = sum(1 for r in all_req if r.priority_level == lvl)
            print(f"{count} request(s) with priority level {lvl}.")
        except ValueError:
            print("Invalid priority level.")
    elif choice == '3':
        types   = Counter(r.request_type for r in all_req)
        prios   = Counter(r.priority_level for r in all_req)
        print("Requests by Type:")
        for t, c in types.items():
            print(f"  {t}: {c}")
        print("\nRequests by Priority Level:")
        for p, c in sorted(prios.items()):
            print(f"  {p}: {c}")
    else:
        print("Invalid choice.")

def add_request_action():
    """
    Prompt for and enqueue a new StudentRequest, recording the action
    on the undo stack for later undo/redo.
    """
    sid_input = input("Student ID: ").strip()
    try:
        sid = int(sid_input)
    except ValueError:
        print(f"{Fore.RED}Invalid Student ID format. Must be an integer.{Style.RESET_ALL}")
        return

    if not student_exists(sid):
        print(f"{Fore.RED}Student ID {sid} not found in system. Cannot add request.{Style.RESET_ALL}")
        return

    existing = [r for r in request_queue.list_all() if r.student_id == sid]
    if existing:
        confirm = input(
            f"{Fore.YELLOW}Student {sid} already has {len(existing)} pending request(s). "
            f"Add another? (YES/NO): {Style.RESET_ALL}"
        ).strip().upper()

        if confirm != "YES":
            print(f"{Fore.CYAN}Request not added.{Style.RESET_ALL}")
            return

    rtype = input("Request Type: ").strip()
    if not rtype:
        print(f"{Fore.RED}Request Type cannot be empty.{Style.RESET_ALL}")
        return

    prio_input = input("Priority Level (integer, lower=more urgent): ").strip()
    try:
        prio = int(prio_input)
    except ValueError:
        print(f"{Fore.RED}Invalid priority; must be an integer.{Style.RESET_ALL}")
        return

    details = input("Request Details: ").strip()
    if not details:
        print(f"{Fore.RED}Request Details cannot be empty.{Style.RESET_ALL}")
        return

    req = StudentRequest(sid, rtype, prio, details)
    request_queue.enqueue(req)
    save_requests()  # ✅ <-- persist change
    undo_stack.append(("enqueue", req))
    redo_stack.clear()

    logging.info(f"Enqueued request: {req!r}")
    print(f"{Fore.GREEN}Enqueued: {req}{Style.RESET_ALL}")




def process_request_action():
    """
    Dequeue the highest-priority request, display and log it,
    and record the action on the undo stack for later undo/redo.
    """
    if request_queue.is_empty():
        print(f"{Fore.YELLOW}No pending requests.{Style.RESET_ALL}")
        return

    # 1) Dequeue
    req = request_queue.dequeue()
    save_requests()

    # 2) Display details
    print(f"\n{Fore.MAGENTA}{Style.BRIGHT}Processing Next Request{Style.RESET_ALL}\n")
    print(f"{Fore.CYAN}Student ID     : {Fore.YELLOW}{req.student_id}")
    print(f"{Fore.CYAN}Request Type   : {Fore.YELLOW}{req.request_type}")
    print(f"{Fore.CYAN}Priority Level : {Fore.YELLOW}{req.priority_level}")
    print(f"{Fore.CYAN}Details        : {Fore.YELLOW}{req.request_details}")
    print(f"{Fore.CYAN}Timestamp      : {Fore.YELLOW}{req.timestamp.isoformat()}\n")

    # 3) Show updated count
    remaining = request_queue.size()
    print(f"{Fore.CYAN}Updated requests remaining: {Fore.YELLOW}{remaining}\n")

    # 4) Log to processed_requests.log
    entry = {
        "processed_at": datetime.now(timezone.utc).isoformat(),
        **req.to_dict()
    }
    with open('processed_requests.log', 'a') as f:
        f.write(json.dumps(entry) + "\n")
    logging.info(f"Processed and logged request: {req!r}")

    # 5) Record undo
    undo_stack.append(("dequeue", req))
    redo_stack.clear()

    print(f"{Fore.GREEN}Request processed and logged.{Style.RESET_ALL}\n")


def display_requests_table(requests):
    headers = ["Req ID", "Student ID", "Type", "Priority", "Timestamp"]
    rows = [
        [
            r.request_id,
            r.student_id,
            r.request_type,
            r.priority_level,
            r.timestamp.strftime("%Y-%m-%d %H:%M:%S")
        ]
        for r in requests
    ]
    col_widths = [
        max(len(str(val)) for val in [hdr] + [row[i] for row in rows])
        for i, hdr in enumerate(headers)
    ]

    # Colorized header
    header_parts = [
        f"{Fore.CYAN}{Style.BRIGHT}{headers[i].ljust(col_widths[i])}{Style.RESET_ALL}"
        for i in range(len(headers))
    ]
    print(" | ".join(header_parts))
    print("-+-".join("-" * col_widths[i] for i in range(len(headers))))

    # Colorized rows
    for row in rows:
        prio = row[3]
        color = Fore.RED if prio == 1 else (Fore.YELLOW if prio == 2 else Fore.GREEN)
        line = " | ".join(str(row[i]).ljust(col_widths[i]) for i in range(len(headers)))
        print(f"{color}{line}{Style.RESET_ALL}")



def generate_dummy_requests(n):
    """
    Generate n dummy StudentRequest objects and enqueue them.
    Uses your existing request_queue.
    """
    sample_types = [
        "Password Reset", "Profile Update", "Transcript Request",
        "Course Enrollment", "Grade Appeal", "Fee Waiver"
    ]
    sample_details = {
        "Password Reset": "Forgot password, needs reset link.",
        "Profile Update": "Change of address and phone number.",
        "Transcript Request": "Official transcript for internship.",
        "Course Enrollment": "Add CS101 to my schedule.",
        "Grade Appeal": "Review grade for assignment 3.",
        "Fee Waiver": "Request waiver for late payment fee."
    }

    existing_ids = list(student_tree.keys())
    for _ in range(n):
        sid = random.choice(existing_ids) if existing_ids else random.randint(10000, 99999)
        rtype = random.choice(sample_types)
        prio = random.randint(1, 5)
        details = sample_details[rtype]
        delta_secs = random.randint(0, 7 * 24 * 3600)
        ts = datetime.now(timezone.utc) - timedelta(seconds=delta_secs)

        req = StudentRequest(sid, rtype, prio, details, timestamp=ts)
        request_queue.enqueue(req)

    save_requests()  # ✅ persist to file
    print(f"Enqueued {n} dummy requests.")


def dashboard_summary():
    """
    Dashboard Summary View:
      - Total number of students
      - Full-time vs Part-time counts
      - Most common course enrolled
      - Average courses per student
      - Total pending requests
      - Breakdown of requests by type
    """
    from collections import Counter

    print(f"\n{Fore.MAGENTA}{Style.BRIGHT}📊 Dashboard Summary{Style.RESET_ALL}\n")

    # 1. Total / FT / PT students
    total_students = len(student_tree)
    full_time      = sum(1 for s in student_tree.values() if s.is_full_time)
    part_time      = total_students - full_time

    # 2. Most common course
    all_courses = [course for s in student_tree.values() for course in s.course_list]
    course_counts = Counter(all_courses)
    if course_counts:
        common_course, common_count = course_counts.most_common(1)[0]
    else:
        common_course, common_count = ("N/A", 0)

    # 3. Average courses per student
    avg_courses = (sum(len(s.course_list) for s in student_tree.values()) /
                   total_students) if total_students else 0.0

    # 4. Pending requests
    pending_requests = request_queue.size()

    # 5. Breakdown of requests by type
    req_types = Counter(r.request_type for r in request_queue.list_all())

    # — now print everything
    print(f"{Fore.CYAN}Total students:            {Fore.YELLOW}{total_students}")
    print(f"{Fore.CYAN} • Full-time:              {Fore.YELLOW}{full_time}")
    print(f"{Fore.CYAN} • Part-time:              {Fore.YELLOW}{part_time}\n")

    print(f"{Fore.CYAN}Most common course:        {Fore.YELLOW}{common_course} ({common_count})")
    print(f"{Fore.CYAN}Average courses/student:   {Fore.YELLOW}{avg_courses:.2f}\n")

    print(f"{Fore.CYAN}Pending student requests:  {Fore.YELLOW}{pending_requests}\n")

    print(f"{Fore.CYAN}Requests by Type:")
    if req_types:
        for typ, cnt in req_types.items():
            print(f"  {Fore.YELLOW}{typ:<20} {Fore.CYAN}→ {Fore.YELLOW}{cnt}")
    else:
        print(f"  {Fore.YELLOW}None")

    print()  # trailing blank line

def login():
    print("Login to the system")
    username = input("Enter username: ").strip()
    password = input("Enter password: ").strip()

    users = {
        "admin": {"password": hashlib.sha256("admin123".encode()).hexdigest(), "role": "admin"},
        "student": {"password": hashlib.sha256("student123".encode()).hexdigest(), "role": "student"}
    }

    hashed_input = hashlib.sha256(password.encode()).hexdigest()
    if username in users and users[username]["password"] == hashed_input:
        logging.info(f"User '{username}' logged in successfully as {users[username]['role']}.")
        print(f"Welcome, {username} ({users[username]['role']})")
        return users[username]["role"]
    else:
        logging.warning(f"Failed login attempt with username '{username}'.")
        print("Invalid username or password.")
        return None

def ai_course_advisory():
    print("\n🤖 AI Course Advisor (OpenRouter) — type 'exit' to return.\n")

    while True:
        user_input = input("You: ").strip()
        if user_input.lower() in {"exit", "quit"}:
            print("Exiting advisor...")
            break

        headers = {
            "Authorization": f"Bearer {os.getenv('OPENROUTER_API_KEY')}",
            "Content-Type": "application/json",
            "HTTP-Referer": "https://chat.openai.com",
            "X-Title": "CourseAdvisorCLI"
        }

        payload = {
            "model": "mistralai/mistral-7b-instruct",  # ✅ active and free
            "messages": [
                {"role": "system", "content": (
                    "You are a helpful and friendly university course advisor. "
                    "Help students choose suitable university courses based on interests, goals, and workload.")},
                {"role": "user", "content": user_input}
            ]
        }

        print("Thinking...", end="", flush=True)
        time.sleep(1)

        try:
            response = requests.post("https://openrouter.ai/api/v1/chat/completions",
                                     headers=headers, json=payload)

            if response.status_code == 200:
                reply = response.json()["choices"][0]["message"]["content"]
                print(f"\nAdvisor: {reply}\n")
            else:
                print(f"\n⚠️ API Error: {response.status_code} - {response.text}\n")

        except requests.RequestException as e:
            print(f"\n⚠️ Request failed: {e}\n")

def fix_encrypted_emails():
    fixed = 0
    for student in student_tree.values():
        if "<decrypt_error>" in student.email:
            try:
                # email is invalid, probably not encrypted yet
                raw_email = getattr(student, '__dict__', {}).get('email')
                if raw_email:
                    student.email = raw_email  # triggers encryption setter
                    del student.__dict__['email']  # cleanup
                    fixed += 1
            except Exception:
                continue
    if fixed > 0:
        print(f"🔧 Fixed {fixed} students with unencrypted email.")
        save_data()
    else:
        print("✅ All emails already encrypted properly.")

def load_student_df():
    with open("student_data.pkl", "rb") as f:
        bst = pickle.load(f)
    rows = []
    for s in bst.values():
        rows.append({
            "id": s.student_id,
            "name": s.name,
            "email": s.email,
            "year": s.year_of_study,
            "status": "Full-time" if s.is_full_time else "Part-time",
            "courses": list(s.course_list)
        })
    return pd.DataFrame(rows)

def load_requests_df():
    with open("requests_data.json", "r") as f:
        data = json.load(f)
    return pd.DataFrame(data)

# Generate charts + PDF
def export_dashboard_charts_pdf(pdf_name=None):
    df_stu = load_student_df()
    df_req = load_requests_df()

    total = len(df_stu)
    ft = int((df_stu["status"] == "Full-time").sum()) if total else 0
    pt = total - ft

    top_course, top_count, vc_courses = "N/A", 0, pd.Series(dtype=int)
    if total:
        exploded = df_stu.explode("courses")
        vc_courses = exploded["courses"].dropna().value_counts()
        if not vc_courses.empty:
            top_course, top_count = vc_courses.index[0], int(vc_courses.iloc[0])

    pending = len(df_req)

    out = pdf_name or f"dashboard_charts_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.pdf"
    with PdfPages(out) as pdf:
        # Page 1: KPI board
        fig1, ax1 = plt.subplots(figsize=(8.3, 5.8))
        ax1.axis("off")
        ax1.text(
            0.05, 0.95,
            (
                "Dashboard Summary (Charts)\n\n"
                f"Total students: {total}\n"
                f"  • Full-time:  {ft}\n"
                f"  • Part-time:  {pt}\n\n"
                f"Most common course: {top_course} ({top_count})\n\n"
                f"Pending requests: {pending}"
            ),
            va="top", ha="left", fontsize=14
        )
        pdf.savefig(fig1); plt.close(fig1)

        # Page 2: FT vs PT bar
        if total:
            fig2, ax2 = plt.subplots(figsize=(8.3, 5.8))
            df_stu["status"].value_counts().sort_index().plot(kind="bar", ax=ax2, rot=0, title="Full-time vs Part-time")
            ax2.set_xlabel(""); ax2.set_ylabel("Count")
            fig2.tight_layout(); pdf.savefig(fig2); plt.close(fig2)

        # Page 3: Top 10 courses
        if total and not vc_courses.empty:
            fig3, ax3 = plt.subplots(figsize=(8.3, 5.8))
            vc_courses.head(10).sort_values(ascending=True).plot(kind="barh", ax=ax3, title="Top 10 Courses by Enrollment")
            ax3.set_xlabel("Students")
            fig3.tight_layout(); pdf.savefig(fig3); plt.close(fig3)

        # Page 4: Requests by Type
        if pending:
            fig4, ax4 = plt.subplots(figsize=(8.3, 5.8))
            df_req["request_type"].value_counts().sort_values(ascending=True).plot(kind="barh", ax=ax4, title="Pending Requests by Type")
            ax4.set_xlabel("Count")
            fig4.tight_layout(); pdf.savefig(fig4); plt.close(fig4)

        # Page 5: Requests by Priority
        if pending:
            fig5, ax5 = plt.subplots(figsize=(8.3, 5.8))
            df_req["priority_level"].value_counts().sort_index().plot(kind="bar", ax=ax5, rot=0, title="Pending Requests by Priority")
            ax5.set_xlabel("Priority"); ax5.set_ylabel("Count")
            fig5.tight_layout(); pdf.savefig(fig5); plt.close(fig5)

    print(f"✅ Exported dashboard charts to '{out}'")


def user(role):
    while True:
        print("\nStudent Course Registration System")

        if role == "admin":
            print(" 1. Display All Students")
            print(" 2. Add New Student")
            print(" 3. Search Student by ID or Name")
            print(" 4. Enroll Student in a Course")
            print(" 5. Remove Student Course")
            print(" 6. Sort by Year of Study (Bubble Sort)")
            print(" 7. Sort by Number of Courses (Selection Sort)")
            print(" 8. Sort by Num of Registered Course & Student ID (Merge Sort)")
            print(" 9. Sort by Year of Study (Quick Sort by Name)")
            print("10. Export to Excel")
            print("11. Import from Excel")
            print("12. Add Student Request to Queue")
            print("13. View Pending Student Requests")
            print("14. View Student Requests Queue Statistics")
            print("15. Process Next Student Request")
            print("16. Undo Last Queue Action")
            print("17. Redo Last Queue Action")
            print("18. Dashboard Summary")
            print("19. Show BST Structure")
            print("20. Show Student Course History")
            print("21. AI Course Advisor")
            print("22. Export Dashboard Charts (PDF)")
            print("23. Logout")
            print("24. Exit")

        elif role == "student":
            print(" 1. Display All Students")
            print(" 2. Search Student by ID or Name")
            print(" 3. View My Course History")
            print(" 4. AI Course Advisor")
            print(" 5. Logout")
            print(" 6. Exit")

        choice = input("Enter your choice: ").strip()

        if role == "admin":
            if choice == '1':
                display_all_students()
            elif choice == '2':
                add_student()
            elif choice == '3':
                search_student()
            elif choice == '4':
                enroll_course()
            elif choice == '5':
                remove_student_course()
            elif choice == '6':
                bubble_sort_year_of_study()
            elif choice == '7':
                selection_sort_num_reg_course()
            elif choice == '8':
                merge_sort_by_courses_and_id()
            elif choice == '9':
                quick_sort_year_name()
            elif choice == '10':
                export_to_excel()
            elif choice == '11':
                import_from_excel()
            elif choice == '12':
                add_request_action()
            elif choice == '13':
                view_requests_menu()
            elif choice == '14':
                view_queue_stats_menu()
            elif choice == '15':
                process_request_action()
            elif choice == '16':
                undo_action()
            elif choice == '17':
                redo_action()
            elif choice == '18':
                dashboard_summary()
            elif choice == '19':
                print("\n── Student BST Structure ──")
                student_tree.print_tree()
            elif choice == '20':
                try:
                    sid = int(input("Enter student ID to view history: ").strip())
                    student = student_tree.search(sid)
                    if student:
                        student.display_history()
                    else:
                        print("Student ID not found.")
                except ValueError:
                    print("Invalid ID format.")
            elif choice == '21':
                ai_course_advisory()
            elif choice == '22':
                custom_pdf = input("PDF name (blank = timestamped): ").strip() or None
                export_dashboard_charts_pdf(pdf_name=custom_pdf)

            elif choice == '23':
                print("Logging out...")
                return
            elif choice == '24':
                print("Exiting program.")
                exit()


        elif role == "student":
            if choice == '1':
                display_all_students()
            elif choice == '2':
                search_student()
            elif choice == '3':
                try:
                    sid = int(input("Enter your student ID: ").strip())
                    student = student_tree.search(sid)
                    if student:
                        student.display_history()
                    else:
                        print("Student ID not found.")
                except ValueError:
                    print("Invalid ID format.")
            elif choice == '4':
                ai_course_advisory()
            elif choice == '5':
                print("Logging out...")
                return
            elif choice == '6':
                print("Exiting program.")
                exit()
            else:
                print("Invalid choice.")



if __name__ == "__main__":
    try:
        load_data()
        load_requests()  # ✅ Load requests
        student = student_tree.search(1001)
        # print(student._encrypted_email)
        # if request_queue.is_empty():
        #     generate_dummy_requests(50)
        fix_encrypted_emails()  # optional
        while True:
            role = login()
            if role:
                user(role)
            else:
                retry = input("Try logging in again? (yes/no): ").strip().lower()
                if retry != 'yes':
                    print("Goodbye!")
                    break
    except Exception as e:
        print("\nExiting Program due to unexpected error:", e)

